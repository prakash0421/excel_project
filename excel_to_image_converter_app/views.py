from PIL import Image, ImageDraw, ImageFont
import openpyxl
from io import BytesIO
from django.shortcuts import render
from django.core.mail import EmailMessage
from django.conf import settings
from .forms import UploadFileForm

def convert_excel_to_image_full(file):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    # Read the data from the Excel sheet
    columns = [cell.value for cell in sheet[1]]
    results = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]

    # Calculate image size
    cell_width = 120
    cell_height = 40
    img_width = cell_width * len(columns)
    img_height = cell_height * (len(results) + 1)  # +1 for headers

    # Create an image with a white background
    img = Image.new('RGB', (img_width, img_height), color='white')
    draw = ImageDraw.Draw(img)

    # Load a font
    font = ImageFont.load_default()

    # Draw the column headers with yellow background
    for i, col_name in enumerate(columns):
        draw.rectangle([i * cell_width, 0, (i + 1) * cell_width, cell_height], fill='yellow')
        draw.text((i * cell_width + 10, 10), str(col_name), fill='black', font=font)

    # Draw the rows with padding
    for row_idx, row in enumerate(results):
        for col_idx, value in enumerate(row):
            draw.text((col_idx * cell_width + 10, (row_idx + 1) * cell_height + 10), str(value), fill='black', font=font)

    # Save the image to a BytesIO object
    image_stream = BytesIO()
    img.save(image_stream, format='JPEG')
    image_stream.seek(0)

    return image_stream

def convert_excel_to_image_specific(file):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    # Ensure the necessary columns are in the Excel file
    columns = [cell.value for cell in sheet[1]]
    required_columns = ['Cust State', 'Cust Pin', 'DPD']
    missing_columns = [col for col in required_columns if col not in columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    # Extract and filter the data
    data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
    pin_index = columns.index('Cust Pin')
    state_index = columns.index('Cust State')

    # Count occurrences of each Cust Pin and store state
    pin_count = {}
    pin_state = {}  # To store state for each pin
    for row in data:
        pin = row[pin_index]
        state = row[state_index]
        if pin not in pin_count:
            pin_count[pin] = 0
            pin_state[pin] = state
        pin_count[pin] += 1

    # Prepare filtered data
    filtered_data = [[pin_state[pin], pin, count] for pin, count in pin_count.items() if count >= 2]
    columns = ['Cust State', 'Cust Pin', 'DPD']  # 'DPD' is the count of occurrences

    # Calculate image size
    cell_width = 120
    cell_height = 40
    img_width = cell_width * len(columns)
    img_height = cell_height * (len(filtered_data) + 1)  # +1 for headers

    # Create an image with a white background
    img = Image.new('RGB', (img_width, img_height), color='white')
    draw = ImageDraw.Draw(img)

    # Load a font
    font = ImageFont.load_default()

    # Draw the column headers with yellow background
    for i, col_name in enumerate(columns):
        draw.rectangle([i * cell_width, 0, (i + 1) * cell_width, cell_height], fill='yellow')
        draw.text((i * cell_width + 10, 10), str(col_name), fill='black', font=font)

    # Draw the rows with padding
    for row_idx, row in enumerate(filtered_data):
        for col_idx, value in enumerate(row):
            draw.text((col_idx * cell_width + 10, (row_idx + 1) * cell_height + 10), str(value), fill='black', font=font)

    # Save the image to a BytesIO object
    image_stream = BytesIO()
    img.save(image_stream, format='JPEG')
    image_stream.seek(0)

    return image_stream


def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            email_to = form.cleaned_data.get('email_to')
            user_name = form.cleaned_data.get('user_name')
            email_body = form.cleaned_data.get('email_body', 'Here is the image generated from the Excel file.')
            image_type = form.cleaned_data.get('image_type', 'any_excel')  # Default to 'any_excel'

            # Construct email subject
            email_subject = f'Python Assignment - {user_name}'

            try:
                # Choose which function to use based on image_type
                if image_type == 'data_set':
                    image_stream = convert_excel_to_image_specific(file)
                else:
                    image_stream = convert_excel_to_image_full(file)

                # Send the image via email
                email = EmailMessage(
                    subject=email_subject,
                    body=email_body,
                    from_email=settings.EMAIL_HOST_USER,
                    to=[email_to],
                )
                image_stream.seek(0)  # Ensure the stream is at the beginning
                email.attach('table_image.jpeg', image_stream.read(), 'image/jpeg')
                email.send()

                return render(request, 'success.html')
            except ValueError as e:
                return render(request, 'upload.html', {'form': form, 'error': str(e)})
            except Exception as e:
                # Log the full exception message
                import traceback
                error_message = str(e)
                error_trace = traceback.format_exc()
                print(error_trace)  # Log to console or file
                return render(request, 'upload.html', {'form': form, 'error': f'An unexpected error occurred: {error_message}'})
    else:
        form = UploadFileForm()

    return render(request, 'upload.html', {'form': form})
